import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
import datetime
import time
from urllib.parse import quote

class ExcelUtil:
    def __init__(self, queryset, field_list=None, header_list=None):
        """
        :param queryset: QuerySet or List of objects/dicts
        :param field_list: List of fields to export (e.g. ['username', 'dept.dept_name'])
        :param header_list: List of headers corresponding to field_list (e.g. ['用户名', '部门'])
        """
        self.queryset = queryset
        self.field_list = field_list or []
        self.header_list = header_list or self.field_list

    def make_excel(self, filename="export"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Headers
        if self.header_list:
            for col_num, header in enumerate(self.header_list, 1):
                cell = ws.cell(row=1, column=col_num, value=str(header))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Set generic width
                ws.column_dimensions[get_column_letter(col_num)].width = 20

        # Data
        for row_num, obj in enumerate(self.queryset, 2):
            for col_num, field in enumerate(self.field_list, 1):
                value = self.get_value(obj, field)
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # output = BytesIO()
        # wb.save(output)
        # output.seek(0)
        # response = HttpResponse(
        #     output.getvalue(),
        #     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # )
        # Ensure filename has extension
        # if not filename.endswith('.xlsx'):
        #    filename += '.xlsx'
        # response.headers['content-disposition'] =  f'attachment;filename={quote(f"{filename}")}'
        # # # cross-origin跨域请求需要设置Access-Control-Expose-Headers响应信息
        # response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        # return response
        return wb

    def get_value(self, obj, field):
        """
        Get value from object, handling nested fields (e.g., 'dept.dept_name')
        and choices display values.
        """
        if isinstance(obj, dict):
            # For dict, we just get the value. Nested dicts support is limited to dot notation manually handled if needed
            # But here we support dot notation for dicts too if keys have dots? No, usually not.
            # Let's support dot notation for dicts too.
            parts = field.split('.')
            value = obj
            try:
                for part in parts:
                    if isinstance(value, dict):
                        value = value.get(part)
                    else:
                        value = getattr(value, part, None)
                    if value is None:
                        break
            except Exception:
                return ''
        else:
            # Model instance
            try:
                # Check for get_FOO_display first if it's a simple field
                if '.' not in field and hasattr(obj, f'get_{field}_display'):
                    return getattr(obj, f'get_{field}_display')()
                
                parts = field.split('.')
                value = obj
                for part in parts:
                    if value is None:
                        break
                    # If current value is a model instance, check for get_field_display again
                    if hasattr(value, f'get_{part}_display'):
                        value = getattr(value, f'get_{part}_display')()
                    else:
                        value = getattr(value, part, None)
            except Exception:
                return ''

        if isinstance(value, datetime.datetime):
            # Remove timezone info for excel compatibility if needed, or just format
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, datetime.date):
            return value.strftime('%Y-%m-%d')
        if value is None:
            return ''
        return str(value)
